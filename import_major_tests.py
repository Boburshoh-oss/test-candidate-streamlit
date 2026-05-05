from __future__ import annotations

import re
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


WORKBOOK_PATH = Path("Junior_Yonalishlar_Savollar_30tadan.xlsx")
MAJOR_DIR = Path("major")
SETTINGS_SHEET = "Yo'nalish sozlamalari"

HEADERS = [
    "No",
    "Savol turi",
    "Savol",
    "A variant",
    "B variant",
    "C variant",
    "D variant",
    "To'g'ri javob",
    "Kutilgan javob/Rubrika",
    "Maks ball",
    "Qiyinlik darajasi",
    "Mavzu",
]

SETTINGS_HEADERS = ["Yo'nalish", "Jami ball", "O'tish foizi", "Izoh"]

SOURCES = [
    {
        "file": "Axborot xavsizligi bo.docx",
        "sheet": "Axborot xavfsizligi",
        "topic": "Axborot xavfsizligi",
        "answers": list("BBCBBBCBBBBBBBA BBC".replace(" ", "")) + ["C", "B", "C"],
    },
    {
        "file": "Dasturiy qo'llab quvvatlash.docx",
        "sheet": "Dasturiy qo'llab-quvvatlash",
        "topic": "Helpdesk Support",
        "answers": None,
    },
    {
        "file": "Sistemniy administratorligidan o.docx",
        "sheet": "System Administrator",
        "topic": "System Administration",
        "answers": [
            "C",
            "B",
            "C",
            "C",
            "C",
            "C",
            "B",
            "C",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
        ],
    },
    {
        "file": "Texnik qo.docx",
        "sheet": "Texnik qo'llab-quvvatlash",
        "topic": "Texnik Support",
        "answers": [
            "A",
            "B",
            "C",
            "B",
            "C",
            "C",
            "B",
            "B",
            "C",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "B",
            "C",
        ],
    },
]


def clean_text(text: str) -> str:
    return (
        text.replace("\u00a0", " ")
        .replace("✔", "")
        .replace("✅", "")
        .replace("☑", "")
        .strip()
    )


def parse_docx(path: Path) -> list[dict[str, str]]:
    doc = Document(path)
    lines: list[str] = []
    for paragraph in doc.paragraphs:
        for line in paragraph.text.splitlines():
            line = line.strip()
            if line:
                lines.append(line)

    questions: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    marked_answer = ""

    for line in lines:
        question_match = re.match(r"^(\d+)\.\s*(.+)$", line)
        option_match = re.match(r"^([A-E])\)\s*(.+)$", line)

        if question_match:
            if current:
                current["marked_answer"] = marked_answer
                questions.append(current)
            current = {"question": clean_text(question_match.group(2))}
            marked_answer = ""
            continue

        if option_match and current is not None:
            letter = option_match.group(1)
            value = option_match.group(2)
            if "✔" in value or "✅" in value or "☑" in value:
                marked_answer = letter
            current[letter] = clean_text(value)

    if current:
        current["marked_answer"] = marked_answer
        questions.append(current)

    return questions


def ensure_headers(ws, headers):
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)


def style_sheet(ws, settings=False):
    fill = PatternFill("solid", fgColor="548235" if settings else "1F4E78")
    font = Font(color="FFFFFF", bold=True)
    side = Side(style="thin", color="D9E2F3")
    border = Border(left=side, right=side, top=side, bottom=side)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def add_validations(ws):
    question_type_validation = DataValidation(type="list", formula1='"test,yozma"', allow_blank=False)
    answer_validation = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=True)
    difficulty_validation = DataValidation(type="list", formula1='"Oson,O\'rta,Qiyin"', allow_blank=True)
    ws.add_data_validation(question_type_validation)
    ws.add_data_validation(answer_validation)
    ws.add_data_validation(difficulty_validation)
    question_type_validation.add("B2:B500")
    answer_validation.add("H2:H500")
    difficulty_validation.add("K2:K500")


def replace_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


def upsert_settings(wb, sheet_name: str, total: int):
    if SETTINGS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SETTINGS_SHEET, 0)
        ws.append(SETTINGS_HEADERS)
    ws = wb[SETTINGS_SHEET]

    existing_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == sheet_name:
            existing_row = row
            break

    row = existing_row or ws.max_row + 1
    values = [sheet_name, total, 60, "Major folderdan import qilingan 20 ta test savol"]
    for column, value in enumerate(values, start=1):
        ws.cell(row=row, column=column, value=value)

    style_sheet(ws, settings=True)
    for col, width in {"A": 28, "B": 14, "C": 14, "D": 52}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def import_sources():
    wb = load_workbook(WORKBOOK_PATH)
    for source in SOURCES:
        path = MAJOR_DIR / source["file"]
        questions = parse_docx(path)
        if len(questions) != 20:
            raise ValueError(f"{path.name}: expected 20 questions, got {len(questions)}")

        answers = source["answers"]
        ws = replace_sheet(wb, source["sheet"])
        ensure_headers(ws, HEADERS)
        for index, question in enumerate(questions, start=1):
            correct = question["marked_answer"] or (answers[index - 1] if answers else "")
            if correct not in {"A", "B", "C", "D"}:
                raise ValueError(f"{path.name}, question {index}: missing correct answer")
            ws.append(
                [
                    index,
                    "test",
                    question["question"],
                    question.get("A", ""),
                    question.get("B", ""),
                    question.get("C", ""),
                    question.get("D", ""),
                    correct,
                    "",
                    1,
                    "O'rta",
                    source["topic"],
                ]
            )

        style_sheet(ws)
        for col, width in {
            "A": 8,
            "B": 14,
            "C": 58,
            "D": 36,
            "E": 36,
            "F": 36,
            "G": 36,
            "H": 14,
            "I": 45,
            "J": 12,
            "K": 18,
            "L": 24,
        }.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        add_validations(ws)
        upsert_settings(wb, source["sheet"], len(questions))

    wb.save(WORKBOOK_PATH)
    wb.close()


if __name__ == "__main__":
    import_sources()
    print(WORKBOOK_PATH.resolve())
