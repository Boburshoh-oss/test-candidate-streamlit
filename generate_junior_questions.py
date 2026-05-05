from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path("Junior_Yonalishlar_Savollar_30tadan.xlsx")

HEADERS = [
    "No",
    "Savol turi",
    "Savol",
    "A variant",
    "B variant",
    "C variant",
    "D variant",
    "To'g'ri javob",
    "Kutilgan javob/Rubrika",
    "Maks ball",
    "Qiyinlik darajasi",
    "Mavzu",
]

SETTINGS_HEADERS = ["Yo'nalish", "Jami ball", "O'tish foizi", "Izoh"]

DISTRACTOR_POOLS = {
    "ai": [
        "Faqat robotlarni boshqaradigan dastur",
        "Faqat internet qidiruv tizimi",
        "Ma'lumotlarni Excelga qo'lda kiritish jarayoni",
        "Faqat o'yin grafikasi uchun kutubxona",
        "Datasetni zip formatga o'tkazish",
        "Server portini sozlash usuli",
        "CSS ranglarini avtomatik tanlash",
        "Modelni tekshirmasdan productionga chiqarish",
        "Faqat parol shifrlash algoritmi",
        "Brauzer cacheini tozalash jarayoni",
        "Faqat HTML sahifa strukturasi",
        "Git commit xabarini yaratish",
        "Database jadval nomi",
        "Foydalanuvchi interfeysi rang palitrasi",
        "Faqat fayl nomini o'zgartirish",
        "Operatsion tizimni qayta o'rnatish",
        "API endpoint URL manzili",
        "Kod formatlash uslubi",
        "Faqat rasm hajmini kichraytirish",
        "Server loglarini o'chirish",
    ],
    "backend": [
        "Faqat HTML sahifa chizish",
        "CSS ranglarini boshqarish",
        "Brauzer animatsiyasini sozlash",
        "Database parolini foydalanuvchiga ko'rsatish",
        "Har doim barcha ma'lumotni bitta response'da qaytarish",
        "Querylarni string qo'shish orqali xavfsiz qilish",
        "HTTP status code ishlatmaslik",
        "Har bir requestda serverni restart qilish",
        "Frontend komponent nomi",
        "Mobil ekran o'lchami",
        "Git branchni o'chirish usuli",
        "Rasm rangini o'zgartirish",
        "API kalitni public repositoryga yozish",
        "Database schema bilan bog'liq bo'lmagan UI matni",
        "Faqat browser cache",
        "Parolni plain text ko'rinishida saqlash",
        "Server xatosini foydalanuvchiga stack trace bilan ko'rsatish",
        "Barcha requestlarni GET bilan yuborish",
        "Input validationni butunlay o'tkazib yuborish",
        "Dependencylarni loyiha bilan bog'lamaslik",
    ],
    "frontend": [
        "Database jadvalini yaratish",
        "Serverdagi migration yozish",
        "API secret keyni HTMLga qo'yish",
        "Brauzerda SQL server ishga tushirish",
        "Har doim fixed width ishlatish",
        "Foydalanuvchi xatosini yashirib qo'yish",
        "Barcha elementlarni bitta divga yozish",
        "Keyboard navigationni e'tiborsiz qoldirish",
        "Loading holatini ko'rsatmaslik",
        "Rasm hajmini nazoratsiz yuklash",
        "Componentlarni qayta ishlatmaslik",
        "State o'rniga global o'zgaruvchilarni tartibsiz ishlatish",
        "CSS selectorlarni keraksiz murakkablashtirish",
        "Button o'rniga faqat oddiy span ishlatish",
        "API xatolarini console'da qoldirish",
        "Mobile viewportni tekshirmaslik",
        "Labelsiz input ishlatish",
        "Noto'g'ri HTML teglar bilan forma yaratish",
        "Har renderda keraksiz request yuborish",
        "Dependencylarni sababsiz ko'paytirish",
    ],
    "qa": [
        "Faqat ranglarni tanlash",
        "Bugni reproduksiya qilmasdan yopish",
        "Expected result yozmaslik",
        "Actual resultni yashirish",
        "Screenshot yoki logni hech qachon qo'shmaslik",
        "Severity va priorityni bir xil deb qabul qilish",
        "Faqat ideal holatlarni tekshirish",
        "Negative testlarni o'tkazib yuborish",
        "Environmentni ko'rsatmaslik",
        "Test data haqida yozmaslik",
        "Bug reportda faqat 'ishlamayapti' deb yozish",
        "API response body'ni tekshirmaslik",
        "Status codelarni e'tiborsiz qoldirish",
        "Regression testni release'dan keyin o'ylash",
        "Acceptance criteriani o'qimaslik",
        "Chegara qiymatlarni tekshirmaslik",
        "Retest qilmasdan bugni tasdiqlash",
        "Flaky testlarni doim ignore qilish",
        "Developer bilan dalilsiz tortishish",
        "Test case qadamlarini tartibsiz yozish",
    ],
    "mobile": [
        "Database jadvalini bevosita UI ichida yaratish",
        "Har ekranga fixed pixel berish",
        "Controllerlarni dispose qilmaslik",
        "Loading holatini ko'rsatmaslik",
        "API xatolarini foydalanuvchidan yashirish",
        "pubspec.yaml indentatsiyasini e'tiborsiz qoldirish",
        "State o'zgarishini UIga bildirmaslik",
        "Scroll kerak joyda scroll bermaslik",
        "Route stackni nazoratsiz ko'paytirish",
        "TextField qiymatini o'qimaslik",
        "Async natijani kutmasdan ishlatish",
        "Assetni e'lon qilmasdan chaqirish",
        "Barcha widgetlarni bitta faylda tartibsiz yozish",
        "Turli ekranlarda tekshirmaslik",
        "Memory leak ehtimolini e'tiborsiz qoldirish",
        "setState'ni disposed widgetda chaqirish",
        "API keyni client kodga ochiq yozish",
        "Form validation qilmaslik",
        "ListView o'rniga overflow qiladigan Column ishlatish",
        "Hot reloadni production deploy deb o'ylash",
    ],
}

CATEGORY_KEYWORDS = {
    "ai": ["AI", "ML", "Model", "Data", "NLP", "Computer Vision", "Generative", "Ethics"],
    "backend": ["Python", "FastAPI", "HTTP", "API", "Database", "Security", "Config", "Git", "PHP", "Laravel", "Validation", "Web", "Architecture"],
    "frontend": ["HTML", "CSS", "JavaScript", "React", "UX", "Forms", "Performance"],
    "qa": ["QA", "Test", "Bug", "API testing", "Tools", "Planning", "Requirement", "Automation", "Teamwork"],
    "mobile": ["Flutter", "Dart", "Layout", "State", "Navigation", "Lifecycle", "Async UI", "Forms"],
}


def category_for(topic: str) -> str:
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(keyword.lower() in topic.lower() for keyword in keywords):
            return category
    return "backend"


def stable_index(*parts: str) -> int:
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()
    return int(digest[:12], 16)


def lengthen_distractor(candidate: str, correct: str, category: str, seed: str) -> str:
    if len(candidate) >= max(20, int(len(correct) * 0.72)):
        return candidate

    tails = {
        "ai": [
            "bo'lib, model baholash yoki data tayyorlash bilan bevosita bog'liq emas",
            "bo'lib, ma'lumotlardan pattern o'rganish jarayonini ifodalamaydi",
            "bo'lib, AI modelning real vazifasini to'g'ri tushuntirib bermaydi",
        ],
        "backend": [
            "bo'lib, backend request-response logikasi bilan bevosita bog'liq emas",
            "bo'lib, API yoki database bilan ishlash vazifasini to'g'ri ifodalamaydi",
            "bo'lib, server tomondagi biznes logikani tushuntirib bermaydi",
        ],
        "frontend": [
            "bo'lib, foydalanuvchi interfeysi holatini boshqarishni to'g'ri ifodalamaydi",
            "bo'lib, browserdagi UI va interaction vazifasiga bevosita mos kelmaydi",
            "bo'lib, responsive yoki component yondashuvini tushuntirib bermaydi",
        ],
        "qa": [
            "bo'lib, test qadamlarini va kutilgan natijani aniq ifodalamaydi",
            "bo'lib, bugni qayta tiklash yoki sifatni tekshirish jarayoniga mos emas",
            "bo'lib, test case yoki bug report mazmunini to'g'ri tushuntirmaydi",
        ],
        "mobile": [
            "bo'lib, Flutter widget daraxti yoki state boshqaruvi bilan mos kelmaydi",
            "bo'lib, mobil UI lifecycle yoki layout vazifasini to'g'ri ifodalamaydi",
            "bo'lib, Dart async yoki Flutter navigation jarayonini tushuntirmaydi",
        ],
    }
    tail = tails[category][stable_index(seed, candidate, correct) % len(tails[category])]
    return f"{candidate} {tail}"


def choose_distractors(question: str, correct: str, topic: str, provided: tuple[str, str, str]) -> list[str]:
    category = category_for(topic)
    pool = [*DISTRACTOR_POOLS[category], *provided]
    start = stable_index(question, topic, correct) % len(pool)
    selected: list[str] = []
    for offset in range(len(pool)):
        candidate = pool[(start + offset) % len(pool)]
        balanced = lengthen_distractor(candidate, correct, category, question)
        if balanced != correct and balanced not in selected:
            selected.append(balanced)
        if len(selected) == 3:
            break
    if len(selected) != 3:
        raise ValueError(f"Not enough unique distractors for: {question}")
    return selected


def assign_options(question: str, correct: str, topic: str, distractors: list[str]) -> tuple[dict[str, str], str]:
    letters = ["A", "B", "C", "D"]
    correct_letter = letters[stable_index("answer", question, topic) % 4]
    options: dict[str, str] = {}
    wrong_iter = iter(distractors)
    for letter in letters:
        options[letter] = correct if letter == correct_letter else next(wrong_iter)
    return options, correct_letter


def test(question, correct, topic, difficulty, wrong1, wrong2, wrong3):
    correct = lengthen_correct_answer(correct, topic)
    distractors = choose_distractors(question, correct, topic, (wrong1, wrong2, wrong3))
    options, correct_letter = assign_options(question, correct, topic, distractors)
    return {
        "type": "test",
        "question": question,
        "a": options["A"],
        "b": options["B"],
        "c": options["C"],
        "d": options["D"],
        "correct": correct_letter,
        "rubric": "",
        "score": 1,
        "difficulty": difficulty,
        "topic": topic,
    }


def lengthen_correct_answer(correct: str, topic: str) -> str:
    if len(correct) >= 22:
        return correct
    additions = {
        "NLP": "bilan ishlaydigan matn va til ma'lumotlari",
        "Computer Vision": "bilan ishlaydigan rasm va video ma'lumotlari",
        "HTTP": f"{correct} uchun ishlatiladigan HTTP amaliyoti",
        "CSS": f"{correct} ko'rinishidagi CSS selector yozilishi",
        "Git": f"{correct} uchun ishlatiladigan versiya nazorati tushunchasi",
        "Dart": f"{correct} sifatida ishlatiladigan Dart tushunchasi",
    }
    for key, value in additions.items():
        if key.lower() in topic.lower():
            return value
    return f"{correct} deb ataladigan asosiy tushuncha"


def written(question, rubric, topic, difficulty):
    return {
        "type": "yozma",
        "question": question,
        "a": "",
        "b": "",
        "c": "",
        "d": "",
        "correct": "",
        "rubric": rubric,
        "score": 3,
        "difficulty": difficulty,
        "topic": topic,
    }


GENERIC_WRONGS = {
    "ai": ("Faqat fayl nomi", "Faqat server sozlamasi", "Faqat UI rangi"),
    "backend": ("Faqat HTML tegi", "Faqat CSS rangi", "Faqat kompyuter nomi"),
    "frontend": ("Database migration", "Server operatsion tizimi", "API kalit"),
    "qa": ("Faqat dizayn rangi", "Database paroli", "Git branch nomi"),
    "mobile": ("Database jadvali", "Server paroli", "CSS class"),
}


def build_ai():
    rows = [
        test("AI nima?", "Kompyuter tizimlarining o'rganish va qaror qabul qilishga o'xshash vazifalarni bajarishi", "AI asoslari", "Oson", *GENERIC_WRONGS["ai"]),
        test("Machine Learning nima?", "Ma'lumotlardan pattern o'rganadigan AI yo'nalishi", "ML asoslari", "Oson", *GENERIC_WRONGS["ai"]),
        test("Supervised learningda datasetda nima bo'ladi?", "Label mavjud bo'ladi", "ML asoslari", "Oson", *GENERIC_WRONGS["ai"]),
        test("Classification vazifasiga qaysi misol mos?", "Email spam yoki spam emasligini aniqlash", "ML vazifalari", "Oson", *GENERIC_WRONGS["ai"]),
        test("Regression vazifasiga qaysi misol mos?", "Narx yoki harorat kabi sonli qiymatni bashorat qilish", "ML vazifalari", "Oson", *GENERIC_WRONGS["ai"]),
        test("Train dataset nima uchun ishlatiladi?", "Modelni o'qitish uchun", "Model training", "Oson", *GENERIC_WRONGS["ai"]),
        test("Test dataset nima uchun kerak?", "Modelni yangi ko'rmagan datada tekshirish uchun", "Model baholash", "Oson", *GENERIC_WRONGS["ai"]),
        test("Accuracy nimani ko'rsatadi?", "To'g'ri bashoratlar ulushini", "Model baholash", "Oson", *GENERIC_WRONGS["ai"]),
        test("Overfitting nima?", "Model train datada yaxshi, yangi datada yomon ishlashi", "Model baholash", "Oson", *GENERIC_WRONGS["ai"]),
        test("Feature nima?", "Model inputidagi alohida belgi yoki ustun", "Data", "Oson", *GENERIC_WRONGS["ai"]),
        test("Label nima?", "Model bashorat qilishi kerak bo'lgan to'g'ri javob", "Data", "Oson", *GENERIC_WRONGS["ai"]),
        test("NLP qaysi data turi bilan ishlaydi?", "Matn va til", "NLP", "Oson", *GENERIC_WRONGS["ai"]),
        test("Computer Vision qaysi data bilan ishlaydi?", "Rasm va video", "Computer Vision", "Oson", *GENERIC_WRONGS["ai"]),
        test("Prompt nima?", "Modelga beriladigan ko'rsatma yoki savol", "Generative AI", "Oson", *GENERIC_WRONGS["ai"]),
        test("Tokenization nima?", "Matnni tokenlarga ajratish", "NLP", "Oson", *GENERIC_WRONGS["ai"]),
        test("Underfitting nima?", "Model train datani ham yaxshi o'rgana olmasligi", "Model baholash", "Oson", *GENERIC_WRONGS["ai"]),
        test("Confusion matrix nima uchun ishlatiladi?", "Classification natijalarini sinflar bo'yicha ko'rish uchun", "Model baholash", "O'rta", *GENERIC_WRONGS["ai"]),
        test("Precision nimaga e'tibor beradi?", "Positive deb topilganlar ichida nechta haqiqatan positive ekaniga", "Model baholash", "O'rta", *GENERIC_WRONGS["ai"]),
        test("Recall nimaga e'tibor beradi?", "Haqiqiy positive holatlarning nechasi topilganiga", "Model baholash", "O'rta", *GENERIC_WRONGS["ai"]),
        test("Data leakage nima?", "Test yoki kelajak ma'lumoti train jarayoniga aralashib ketishi", "Data", "O'rta", *GENERIC_WRONGS["ai"]),
        written("Supervised learning va unsupervised learning farqini misol bilan tushuntiring.", "Supervised learningda label bo'ladi va model inputdan to'g'ri javobni o'rganadi; misol spam/spam emas. Unsupervised learningda label bo'lmaydi, model pattern yoki klaster topadi; misol mijozlarni guruhlash.", "ML asoslari", "Oson"),
        written("Overfitting nima va uni kamaytirish uchun nimalarga e'tibor beriladi?", "Overfitting train datani yodlab olish va testda yomon ishlashdir. Yechimlar: ko'proq data, train/test split, cross-validation, regularization, modelni soddalashtirish, early stopping.", "Model baholash", "Oson"),
        written("Accuracy doim yetarlimi? Qisqa tushuntiring.", "Accuracy ba'zi hollarda yetarli, lekin imbalance datasetda noto'g'ri taassurot beradi. Precision, recall, F1 va confusion matrix ham kerak bo'lishi mumkin.", "Model baholash", "O'rta"),
        written("Prompt yozishda aniq ko'rsatma berish nima uchun muhim?", "Model vazifani, formatni, cheklovlarni va kutilgan natijani yaxshiroq tushunadi. Aniq prompt xatolarni kamaytiradi va javob sifatini oshiradi.", "Generative AI", "Oson"),
        written("Train, validation va test dataset rollarini tushuntiring.", "Train modelni o'qitadi, validation sozlash va tanlashda yordam beradi, test esa yakuniy baholash uchun model ko'rmagan data sifatida ishlatiladi.", "Model training", "O'rta"),
        written("Precision va recall farqini oddiy misolda tushuntiring.", "Precision positive deb topilganlar orasidagi to'g'ri positive ulushi, recall esa haqiqiy positive holatlarning topilgan ulushi. Kasallik aniqlashda recall muhim bo'lishi mumkin.", "Model baholash", "O'rta"),
        written("Junior AI specialist datasetni tayyorlashda qanday tekshiruvlar qiladi?", "Missing values, duplicate qatorlar, noto'g'ri formatlar, outlierlar, label sifati, train/test ajratish va data leakage xavfi tekshiriladi.", "Data", "O'rta"),
        written("NLP vazifasiga bitta real misol keltiring va qanday data kerakligini yozing.", "Misol: review sentiment analysis, chatbot, matn klassifikatsiyasi. Matnlar va kerak bo'lsa label, tozalangan dataset, train/test ajratish kerak.", "NLP", "Oson"),
        written("Qachon model juda sodda bo'lib qoladi va bu qanday bilinadi?", "Underfittingda model train va testda ham yomon ishlaydi. Belgilar: past accuracy, yuqori error, patternlarni o'rgana olmaslik. Model yoki featurelar yetarli emas bo'lishi mumkin.", "Model baholash", "Qiyin"),
        written("AI loyihada etik masalalarga junior darajada nimalar kiradi?", "Bias, shaxsiy ma'lumotlarni himoya qilish, noto'g'ri qarorlar ta'siri, data manbasi, model cheklovlarini tushuntirish va inson nazorati.", "AI ethics", "Qiyin"),
    ]
    return rows


def build_backend_python():
    w = GENERIC_WRONGS["backend"]
    return [
        test("Python'da list nima?", "Tartibli va o'zgaruvchan kolleksiya", "Python asoslari", "Oson", *w),
        test("Dictionary qanday ma'lumot saqlaydi?", "Key-value juftliklar", "Python asoslari", "Oson", *w),
        test("Virtual environment nima uchun kerak?", "Loyiha dependencylarini alohida saqlash uchun", "Muhit", "Oson", *w),
        test("pip nima?", "Python paketlarini o'rnatish vositasi", "Muhit", "Oson", *w),
        test("FastAPI asosan nima uchun ishlatiladi?", "Python API yaratish uchun", "FastAPI", "Oson", *w),
        test("HTTP GET odatda nima uchun ishlatiladi?", "Ma'lumot olish", "HTTP", "Oson", *w),
        test("HTTP POST odatda nima uchun ishlatiladi?", "Yangi ma'lumot yuborish yoki yaratish", "HTTP", "Oson", *w),
        test("JSON nima?", "Ma'lumot almashish formati", "API", "Oson", *w),
        test("Status code 404 nimani bildiradi?", "Resurs topilmadi", "HTTP", "Oson", *w),
        test("Status code 200 nimani bildiradi?", "So'rov muvaffaqiyatli", "HTTP", "Oson", *w),
        test("SQL database nima uchun ishlatiladi?", "Strukturali ma'lumotlarni saqlash va so'rash uchun", "Database", "Oson", *w),
        test("ORM nima?", "Kod orqali database jadvali bilan ishlashni osonlashtiruvchi qatlam", "Database", "Oson", *w),
        test("Environment variable nima uchun ishlatiladi?", "Maxfiy yoki muhitga bog'liq sozlamalarni saqlash uchun", "Config", "Oson", *w),
        test("Git nima?", "Versiya nazorati tizimi", "Git", "Oson", *w),
        test("API endpoint nima?", "Client murojaat qiladigan URL manzil", "API", "Oson", *w),
        test("Exception handling nima uchun kerak?", "Xatolarni boshqarish uchun", "Python asoslari", "Oson", *w),
        test("Pydantic FastAPI'da nima uchun ishlatiladi?", "Data validation va schema uchun", "FastAPI", "O'rta", *w),
        test("Authentication va authorization farqi nimada?", "Authentication kimligini, authorization ruxsatni tekshiradi", "Security", "O'rta", *w),
        test("SQL injection nima?", "Zararli SQL kiritish hujumi", "Security", "O'rta", *w),
        test("Pagination nima uchun kerak?", "Katta ro'yxatni bo'lib-bo'lib qaytarish uchun", "API", "O'rta", *w),
        written("REST API endpoint loyihalashda nimalarga e'tibor beriladi?", "HTTP metodlar to'g'ri tanlanishi, URL nomlari tushunarli bo'lishi, status code, request/response JSON, validation, auth va error format izchil bo'lishi kerak.", "API", "Oson"),
        written("Python'da list va tuple farqini tushuntiring.", "List o'zgaruvchan, tuple o'zgarmas. List element qo'shish/o'chirish uchun qulay, tuple esa doimiy qiymatlar uchun ishlatiladi.", "Python asoslari", "Oson"),
        written("FastAPI'da request body validation nima uchun kerak?", "Client yuborgan data kutilgan formatda ekanini tekshiradi. Pydantic schema noto'g'ri data biznes logikaga o'tib ketishini kamaytiradi.", "FastAPI", "O'rta"),
        written("Backend loyihada .env faylga nimalar qo'yiladi?", "API key, database URL, secret key kabi muhit sozlamalari qo'yiladi. Public repo'ga commit qilmaslik, example fayl yaratish va productionda xavfsiz saqlash kerak.", "Config", "Oson"),
        written("Database migration nima va nima uchun kerak?", "Database schema o'zgarishlarini versiyalash jarayoni. Jadval qo'shish, ustun o'zgartirish kabi ishlarni jamoada izchil bajarishga yordam beradi.", "Database", "O'rta"),
        written("API error response qanday bo'lishi yaxshi?", "Status code to'g'ri, xabar tushunarli, kerakli fieldlar ko'rsatilgan, maxfiy ichki xatolar oshkor qilinmagan va format izchil bo'lishi kerak.", "API", "O'rta"),
        written("SQL injectiondan himoyalanishning asosiy yo'llarini yozing.", "Parameterized queries, ORM, input validation, string concatenation bilan query tuzmaslik, minimal database ruxsatlari va xatolarni oshkor qilmaslik.", "Security", "O'rta"),
        written("Unit test backendda nima uchun kerak?", "Funksiya yoki kichik logika qismlari kutilganidek ishlashini tekshiradi. Refactor va bug fixlarda regressiyani kamaytiradi.", "Testing", "Oson"),
        written("JWT nima va backendda qayerda ishlatiladi?", "JWT token foydalanuvchi identity/claimlarini saqlash va stateless authenticationda ishlatilishi mumkin. Token imzolangan bo'ladi, maxfiy data ko'p qo'yilmasligi kerak.", "Security", "Qiyin"),
        written("Async endpoint qachon foydali bo'lishi mumkin?", "Ko'p I/O kutish bor bo'lsa, masalan tashqi API, database, fayl o'qish. CPU-heavy vazifalar uchun async har doim yechim emas.", "FastAPI", "Qiyin"),
    ]


def build_backend_php():
    w = GENERIC_WRONGS["backend"]
    return [
        test("PHP asosan qaysi sohada ishlatiladi?", "Backend web dasturlash", "PHP asoslari", "Oson", *w),
        test("PHP fayllar odatda qaysi kengaytmaga ega?", ".php", "PHP asoslari", "Oson", *w),
        test("Composer nima?", "PHP dependency manager", "Muhit", "Oson", *w),
        test("Laravel nima?", "PHP web framework", "Laravel", "Oson", *w),
        test("MVC nima?", "Model, View, Controller", "Architecture", "Oson", *w),
        test("HTTP GET odatda nima uchun ishlatiladi?", "Ma'lumot olish", "HTTP", "Oson", *w),
        test("HTTP POST odatda nima uchun ishlatiladi?", "Ma'lumot yuborish yoki yaratish", "HTTP", "Oson", *w),
        test("Status code 500 nimani bildiradi?", "Server xatosi", "HTTP", "Oson", *w),
        test("Associative array nima?", "Key-value juftlik saqlaydigan array", "PHP asoslari", "Oson", *w),
        test("PDO nima uchun ishlatiladi?", "PHP'da database bilan xavfsiz ishlash uchun", "Database", "Oson", *w),
        test("SQL nima?", "Database bilan so'rovlar tili", "Database", "Oson", *w),
        test("Middleware nima?", "Request va response oralig'ida ishlaydigan qatlam", "Laravel", "Oson", *w),
        test("Route nima?", "URL va controller/action bog'lanishi", "Laravel", "Oson", *w),
        test("Validation nima uchun kerak?", "Kiritilgan data to'g'ri formatda ekanini tekshirish", "Validation", "Oson", *w),
        test("Session nima uchun ishlatiladi?", "Foydalanuvchi holatini server tomonda saqlash uchun", "Web", "Oson", *w),
        test(".env faylda nimalar bo'lishi mumkin?", "DB parol, app key, muhit sozlamalari", "Config", "Oson", *w),
        test("CSRF himoyasi nima uchun kerak?", "Foydalanuvchi nomidan ruxsatsiz request yuborishni kamaytirish uchun", "Security", "O'rta", *w),
        test("Eloquent nima?", "Laravel ORM", "Laravel", "O'rta", *w),
        test("SQL injection nimadan kelib chiqadi?", "User inputni xavfsiz parametrsiz queryga qo'shishdan", "Security", "O'rta", *w),
        test("Migration nima?", "Database schema o'zgarishini versiyalash", "Database", "O'rta", *w),
        written("Laravel MVC oqimini junior darajada tushuntiring.", "Route requestni qabul qiladi, controller biznes logika yoki service chaqiradi, model database bilan ishlaydi, view esa UI qaytaradi. API bo'lsa JSON response qaytadi.", "Laravel", "Oson"),
        written("PHP associative array va indexed array farqini yozing.", "Indexed array elementlarga raqamli index orqali murojaat qiladi. Associative array key-value ko'rinishida bo'ladi, masalan user['name'].", "PHP asoslari", "Oson"),
        written("Form validationda nimalarni tekshirish kerak?", "Majburiy fieldlar, data turi, uzunlik, email/telefon formati, unique qiymat, xavfsiz belgilar va biznes qoida talablarini tekshirish kerak.", "Validation", "Oson"),
        written("SQL injectiondan himoyalanish uchun PHP/Laravelda qanday yo'llar bor?", "Prepared statements, PDO binding, Eloquent/query builder, validation, raw querylarni ehtiyot ishlatish va errorlarni oshkor qilmaslik.", "Security", "O'rta"),
        written("Composer nima uchun jamoaviy loyihalarda muhim?", "Dependency versiyalarini boshqaradi, composer.json/lock orqali hamma bir xil paketlarni o'rnatadi, autoload beradi va package managementni osonlashtiradi.", "Muhit", "Oson"),
        written("Laravel middleware uchun real misol keltiring.", "Authentication tekshirish, role permission, request logging, locale o'rnatish, rate limiting kabi request controllerga yetishidan oldin ishlaydigan vazifalar.", "Laravel", "O'rta"),
        written("API response formatini izchil qilish nima uchun kerak?", "Frontend va boshqa clientlar response'ni oldindan biladi, error handling osonlashadi, debugging va documentation yaxshilanadi.", "API", "O'rta"),
        written("Database migration va seeder farqini tushuntiring.", "Migration schema yaratadi yoki o'zgartiradi. Seeder esa test yoki boshlang'ich ma'lumotlarni databasega kiritadi.", "Database", "O'rta"),
        written("Authentication va authorization farqini PHP backend misolida tushuntiring.", "Authentication foydalanuvchi kimligini tekshiradi, masalan login. Authorization esa u admin sahifaga kira oladimi yoki postni o'chira oladimi kabi ruxsatni tekshiradi.", "Security", "Qiyin"),
        written("N+1 query muammosi nima?", "Loop ichida har element uchun alohida query ketib qolishi. Laravelda eager loading bilan kamaytiriladi. Junior buni query soni ortib ketishi sifatida tushunishi kerak.", "Database", "Qiyin"),
    ]


def build_frontend():
    w = GENERIC_WRONGS["frontend"]
    return [
        test("HTML nima uchun ishlatiladi?", "Web sahifa strukturasi uchun", "HTML", "Oson", *w),
        test("CSS nima uchun ishlatiladi?", "Web sahifa ko'rinishini bezash uchun", "CSS", "Oson", *w),
        test("JavaScript nima uchun ishlatiladi?", "Web sahifaga interaktivlik qo'shish uchun", "JavaScript", "Oson", *w),
        test("React nima?", "UI yaratish uchun JavaScript kutubxonasi", "React", "Oson", *w),
        test("Component nima?", "Qayta ishlatiladigan UI bo'lagi", "React", "Oson", *w),
        test("Props nima?", "Parent componentdan child componentga uzatiladigan data", "React", "Oson", *w),
        test("State nima?", "Component ichidagi o'zgaruvchan holat", "React", "Oson", *w),
        test("Responsive design nima?", "Turli ekran o'lchamlariga moslashuvchi dizayn", "CSS", "Oson", *w),
        test("Flexbox nima uchun ishlatiladi?", "Elementlarni bir o'q bo'yicha joylashtirish uchun", "CSS", "Oson", *w),
        test("CSS class selector qanday yoziladi?", ".className", "CSS", "Oson", *w),
        test("ID selector qanday yoziladi?", "#idName", "CSS", "Oson", *w),
        test("Button click event qaysi tushunchaga kiradi?", "User interaction event", "JavaScript", "Oson", *w),
        test("API'dan data olishda frontendda ko'pincha nima ishlatiladi?", "fetch yoki axios", "API", "Oson", *w),
        test("Form input validation nima uchun kerak?", "Foydalanuvchi kiritgan data formatini tekshirish uchun", "Forms", "Oson", *w),
        test("Git branch nima uchun ishlatiladi?", "Alohida feature yoki o'zgarish ustida ishlash uchun", "Git", "Oson", *w),
        test("npm nima?", "JavaScript paket manager", "Muhit", "Oson", *w),
        test("useEffect qachon ishlatiladi?", "Side effectlar: data fetch, subscription, document title kabi ishlar uchun", "React", "O'rta", *w),
        test("Controlled input nima?", "Input qiymati React state orqali boshqariladi", "React", "O'rta", *w),
        test("Accessibility nima?", "Ilovani yordamchi texnologiyalar uchun ham qulay qilish", "UX", "O'rta", *w),
        test("CORS xatosi nimaga bog'liq?", "Browserning cross-origin request xavfsizlik siyosatiga", "API", "O'rta", *w),
        written("React component, props va state farqini tushuntiring.", "Component UI bo'lagi, props parentdan keladigan data, state esa component ichida o'zgaradigan holat. State o'zgarsa UI qayta render bo'ladi.", "React", "Oson"),
        written("Responsive layout qilishda nimalarga e'tibor beriladi?", "Mobile-first yondashuv, flexible width, media query, readable text, button o'lchami, overflow bo'lmasligi va turli viewportlarda tekshirish.", "CSS", "Oson"),
        written("Form submit bo'lganda frontend qanday tekshiruvlar qilishi mumkin?", "Required field, email/telefon format, minimal uzunlik, parol mosligi, son oralig'i, xatolik xabarlarini tushunarli ko'rsatish.", "Forms", "Oson"),
        written("API'dan data olishda loading va error state nima uchun kerak?", "Foydalanuvchi kutish holatini ko'radi, xato bo'lsa nima bo'lganini tushunadi. UI bo'sh yoki muzlab qolgandek ko'rinmaydi.", "API", "O'rta"),
        written("Semantic HTML nima va nima uchun foydali?", "Ma'noli teglar ishlatish: header, nav, main, button, form. Accessibility, SEO va kod o'qilishi yaxshilanadi.", "HTML", "Oson"),
        written("CSS specificity nima ekanini oddiy tushuntiring.", "Qaysi CSS qoida ustun kelishini belgilaydi. Inline, ID, class, tag selectorlarning og'irligi turlicha. Keraksiz murakkab selectorlardan qochish kerak.", "CSS", "O'rta"),
        written("Frontendda accessibility uchun 4 ta oddiy amaliyot yozing.", "Buttonni button sifatida ishlatish, label qo'yish, alt text, keyboard focus, rang kontrasti, aria faqat kerak bo'lsa ishlatish.", "UX", "O'rta"),
        written("React list render qilganda key nima uchun kerak?", "React elementlarni samarali taqqoslash va yangilash uchun key ishlatadi. Stable unique key noto'g'ri render va performance muammolarini kamaytiradi.", "React", "O'rta"),
        written("CORS muammosida frontend va backend roli qanday?", "Frontend cross-origin request yuboradi, lekin ruxsatni backend headerlar orqali beradi. Muammo ko'pincha backend CORS sozlamasi bilan hal qilinadi.", "API", "Qiyin"),
        written("Junior frontend developer performance uchun nimalarga e'tibor berishi mumkin?", "Keraksiz renderlarni kamaytirish, rasm hajmini optimallashtirish, lazy loading, katta dependencylardan ehtiyot bo'lish, pagination va network requestlarni nazorat qilish.", "Performance", "Qiyin"),
    ]


def build_qa():
    w = GENERIC_WRONGS["qa"]
    return [
        test("QA nima?", "Software sifatini tekshirish va yaxshilash jarayoni", "QA asoslari", "Oson", *w),
        test("Bug nima?", "Kutilgan natijaga zid xato yoki nuqson", "QA asoslari", "Oson", *w),
        test("Test case nima?", "Aniq shart, qadam va kutilgan natijaga ega tekshiruv", "Test design", "Oson", *w),
        test("Expected result nimani bildiradi?", "Kutilgan natija", "Test design", "Oson", *w),
        test("Actual result nimani bildiradi?", "Test paytida olingan real natija", "Test design", "Oson", *w),
        test("Regression testing nima?", "Yangi o'zgarish eski funksiyani buzmaganini tekshirish", "Testing turlari", "Oson", *w),
        test("Smoke testing nima?", "Asosiy funksiyalar umuman ishlayotganini tez tekshirish", "Testing turlari", "Oson", *w),
        test("Sanity testing nima?", "Kichik o'zgarishdan keyin tegishli joy ishlashini tekshirish", "Testing turlari", "Oson", *w),
        test("Severity nimani bildiradi?", "Bug tizimga qanchalik jiddiy ta'sir qilishini", "Bug report", "Oson", *w),
        test("Priority nimani bildiradi?", "Bug qanchalik tez tuzatilishi kerakligini", "Bug report", "Oson", *w),
        test("Bug reportda nimalar bo'lishi kerak?", "Title, qadamlar, expected, actual, environment, screenshot/log", "Bug report", "Oson", *w),
        test("API testingda nimani tekshiramiz?", "Status code, response body, validation, auth, error holatlar", "API testing", "Oson", *w),
        test("Postman nima uchun ishlatiladi?", "API so'rovlarini test qilish uchun", "Tools", "Oson", *w),
        test("Boundary value testing nima?", "Chegara qiymatlarni tekshirish", "Test design", "Oson", *w),
        test("Positive test nima?", "To'g'ri input bilan kutilgan muvaffaqiyatli holatni tekshirish", "Test design", "Oson", *w),
        test("Negative test nima?", "Noto'g'ri input yoki xato holatlarni tekshirish", "Test design", "Oson", *w),
        test("Exploratory testing nima?", "Tester tajriba va kuzatishga asoslanib erkin tekshirishi", "Testing turlari", "O'rta", *w),
        test("Test plan nima?", "Test scope, yondashuv, resurs va jadvalni belgilovchi hujjat", "Planning", "O'rta", *w),
        test("Retest nima?", "Tuzatilgan bug qayta tekshirilishi", "Bug lifecycle", "O'rta", *w),
        test("Acceptance criteria nima?", "Feature qachon tayyor deb qabul qilinishini belgilovchi shartlar", "Requirement", "O'rta", *w),
        written("Yaxshi bug report qanday bo'lishi kerak?", "Aniq title, reproducible qadamlar, expected va actual result, environment, test data, screenshot/video/log, severity/priority taklifi bo'lishi kerak.", "Bug report", "Oson"),
        written("Smoke test va regression test farqini tushuntiring.", "Smoke test asosiy funksiyalar ishlayotganini tez tekshiradi. Regression test yangi o'zgarishlar eski funksiyalarni buzmaganini kengroq tekshiradi.", "Testing turlari", "Oson"),
        written("Login formani test qilish uchun 5 ta test case yozing.", "To'g'ri login, noto'g'ri parol, bo'sh fieldlar, noto'g'ri email format, parol ko'rish/yashirish, lockout yoki validation xabarlarini kiritish mumkin.", "Test design", "Oson"),
        written("Severity va priority farqini misol bilan tushuntiring.", "Severity ta'sir darajasi, priority tuzatish tezligi. Masalan logo xatosi low severity lekin release oldi high priority bo'lishi mumkin.", "Bug report", "O'rta"),
        written("API endpoint testida junior QA nimalarni tekshiradi?", "Status code, response schema, required fieldlar, auth, invalid input, boundary values, error message, response timening oddiy darajasi.", "API testing", "O'rta"),
        written("Boundary value testing uchun oddiy misol keltiring.", "Parol uzunligi 8-20 bo'lsa 7, 8, 20, 21 qiymatlarni tekshirish. Chegaralarda xato ko'p chiqishi mumkin.", "Test design", "Oson"),
        written("Test case yozishda precondition va test data nima uchun kerak?", "Test boshlanishi uchun kerakli holat va data aniqlanadi. Bu testni qayta bajarish, tushunish va natijani solishtirishni osonlashtiradi.", "Test design", "O'rta"),
        written("QA developer bilan bug haqida qanday muloqot qilishi kerak?", "Aniq, dalil bilan, ayblamasdan, reproduksiya qadamlarini berib, kerak bo'lsa log/screenshot ulashib va savollarga ochiq bo'lishi kerak.", "Teamwork", "O'rta"),
        written("Flaky test nima va uni aniqlashda nimalarga qaraysiz?", "Ba'zan o'tib, ba'zan yiqiladigan test. Timing, test data, muhit, external dependency, parallel execution va noto'g'ri kutishlar tekshiriladi.", "Automation", "Qiyin"),
        written("QA automation junior uchun qaysi holatlarni avtomatlashtirishga arziydi?", "Tez-tez takrorlanadigan, barqaror, biznes uchun muhim regression scenariylar. Juda tez o'zgaradigan yoki bir martalik tekshiruvlar qo'lda qolishi mumkin.", "Automation", "Qiyin"),
    ]


def build_mobile_flutter():
    w = GENERIC_WRONGS["mobile"]
    return [
        test("Flutter nima?", "Cross-platform UI framework", "Flutter asoslari", "Oson", *w),
        test("Dart nima?", "Flutterda ishlatiladigan dasturlash tili", "Dart", "Oson", *w),
        test("Widget nima?", "Flutter UI qurilish bloki", "Flutter UI", "Oson", *w),
        test("StatelessWidget nima?", "Ichki holati o'zgarmaydigan widget", "Flutter UI", "Oson", *w),
        test("StatefulWidget nima?", "Holati o'zgarishi mumkin bo'lgan widget", "Flutter UI", "Oson", *w),
        test("setState nima qiladi?", "State o'zgarganini bildirib UI qayta chizilishiga sabab bo'ladi", "State", "Oson", *w),
        test("MaterialApp nima?", "Flutter app uchun root konfiguratsiya", "Flutter UI", "Oson", *w),
        test("Scaffold nima beradi?", "AppBar, body, drawer kabi sahifa strukturasi", "Flutter UI", "Oson", *w),
        test("Column widget nima qiladi?", "Child widgetlarni vertikal joylashtiradi", "Layout", "Oson", *w),
        test("Row widget nima qiladi?", "Child widgetlarni gorizontal joylashtiradi", "Layout", "Oson", *w),
        test("pubspec.yaml nima uchun kerak?", "Dependency, asset va loyiha sozlamalari uchun", "Muhit", "Oson", *w),
        test("Hot reload nima?", "Kod o'zgarishini tez UIga qo'llash", "Muhit", "Oson", *w),
        test("Navigator nima uchun ishlatiladi?", "Sahifalar orasida o'tish uchun", "Navigation", "Oson", *w),
        test("TextEditingController nima uchun kerak?", "TextField qiymatini boshqarish uchun", "Forms", "Oson", *w),
        test("Future nima?", "Kelajakda tugaydigan async natija", "Dart", "Oson", *w),
        test("async/await nima uchun ishlatiladi?", "Asinxron operatsiyalarni o'qilishi oson tarzda kutish uchun", "Dart", "Oson", *w),
        test("ListView nima uchun ishlatiladi?", "Scroll bo'ladigan ro'yxat ko'rsatish uchun", "Flutter UI", "O'rta", *w),
        test("Expanded widget nima qiladi?", "Row/Column ichida qolgan joyni egallashga yordam beradi", "Layout", "O'rta", *w),
        test("initState qachon chaqiriladi?", "State obyekt yaratilganda bir marta", "Lifecycle", "O'rta", *w),
        test("dispose nima uchun kerak?", "Controller yoki subscription kabi resurslarni tozalash uchun", "Lifecycle", "O'rta", *w),
        written("StatelessWidget va StatefulWidget farqini tushuntiring.", "StatelessWidget ichki o'zgaruvchi holatga muhtoj emas. StatefulWidget state saqlaydi va setState orqali UI yangilanadi. Form, counter, loading holatlari Stateful bo'lishi mumkin.", "Flutter UI", "Oson"),
        written("Flutterda Row, Column va Expanded qanday ishlatiladi?", "Row gorizontal, Column vertikal joylashtiradi. Expanded qolgan joyni egallashga yordam beradi. Overflow bo'lmasligi uchun flexible layout va scroll kerak bo'lishi mumkin.", "Layout", "Oson"),
        written("TextField bilan forma yaratishda nimalarga e'tibor beriladi?", "Controller, validation, keyboard type, error text, submit holati, bo'sh fieldlarni tekshirish va controllerlarni dispose qilishga e'tibor beradi.", "Forms", "O'rta"),
        written("API'dan data olishda loading, success va error state nima uchun kerak?", "Foydalanuvchi kutish, muvaffaqiyatli data va xato holatlarini ko'radi. App bo'sh yoki qotib qolgandek ko'rinmaydi.", "API", "O'rta"),
        written("pubspec.yaml faylida dependency va asset qo'shish jarayonini tushuntiring.", "Dependency nomi va versiyasi dependenciesga qo'shiladi, keyin flutter pub get qilinadi. Assetlar assets bo'limida e'lon qilinadi va indentatsiya to'g'ri bo'lishi kerak.", "Muhit", "Oson"),
        written("Navigator bilan sahifalar orasida o'tishning oddiy misolini tushuntiring.", "Navigator.push yangi sahifani stackka qo'shadi, Navigator.pop oldingi sahifaga qaytaradi. Route yoki MaterialPageRoute ishlatilishi mumkin.", "Navigation", "Oson"),
        written("setState'dan noto'g'ri foydalanish qanday muammolar keltirishi mumkin?", "Keraksiz rebuild, performance pasayishi, disposed widgetda setState chaqirish xatosi. State o'zgaradigan minimal joyda ishlatish kerak.", "State", "O'rta"),
        written("Flutter appda responsive UI uchun nimalarga e'tibor beriladi?", "MediaQuery, LayoutBuilder, flexible widgetlar, scroll, text overflow, turli ekranlarda tekshirish, fixed widthlardan ehtiyot bo'lish.", "Layout", "O'rta"),
        written("FutureBuilder nima uchun ishlatiladi?", "Future natijasiga qarab loading, error yoki data UI ko'rsatish uchun. Masalan API chaqiruvi natijasini ko'rsatish.", "Async UI", "Qiyin"),
        written("dispose qilinmasa qanday muammolar bo'lishi mumkin?", "Controller, stream yoki listenerlar xotirada qolishi, memory leak, eski callbacklar ishlashi va xatolar yuzaga kelishi mumkin.", "Lifecycle", "Qiyin"),
    ]


BANKS = {
    "AI": build_ai(),
    "Backend Python": build_backend_python(),
    "Backend PHP": build_backend_php(),
    "Frontend": build_frontend(),
    "QA": build_qa(),
    "Mobile Flutter": build_mobile_flutter(),
}


def validate_banks():
    rebalance_answer_letters()
    for direction, rows in BANKS.items():
        assert len(rows) == 30, (direction, len(rows))
        assert sum(1 for row in rows if row["type"] == "test") == 20, direction
        assert sum(1 for row in rows if row["type"] == "yozma") == 10, direction
        assert sum(float(row["score"]) for row in rows) == 50, direction
        counts = {name: sum(1 for row in rows if row["difficulty"] == name) for name in ("Oson", "O'rta", "Qiyin")}
        assert counts == {"Oson": 20, "O'rta": 8, "Qiyin": 2}, (direction, counts)
        test_rows = [row for row in rows if row["type"] == "test"]
        answer_counts = {letter: sum(1 for row in test_rows if row["correct"] == letter) for letter in ("A", "B", "C", "D")}
        assert answer_counts == {"A": 5, "B": 5, "C": 5, "D": 5}, (direction, answer_counts)
        option_sets = {(row["a"], row["b"], row["c"], row["d"]) for row in test_rows}
        assert len(option_sets) == len(test_rows), direction
        for row in test_rows:
            options = [row["a"], row["b"], row["c"], row["d"]]
            assert len(set(options)) == 4, (direction, row["question"])
            correct_text = row[row["correct"].lower()]
            option_lengths = [len(option) for option in options]
            assert len(correct_text) <= max(option_lengths), (direction, row["question"])
            assert min(option_lengths) >= max(18, int(len(correct_text) * 0.55)), (direction, row["question"], option_lengths)


def rebalance_answer_letters():
    letters = ["A", "B", "C", "D"]
    for rows in BANKS.values():
        test_index = 0
        for row in rows:
            if row["type"] != "test":
                continue
            target_letter = letters[test_index % len(letters)]
            if row["correct"] != target_letter:
                current_letter = row["correct"]
                row[current_letter.lower()], row[target_letter.lower()] = row[target_letter.lower()], row[current_letter.lower()]
                row["correct"] = target_letter
            test_index += 1


def style_header(ws, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    border_side = Side(style="thin", color="D9E2F3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def add_validations(ws):
    question_type_validation = DataValidation(type="list", formula1='"test,yozma"', allow_blank=False)
    answer_validation = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=True)
    difficulty_validation = DataValidation(type="list", formula1='"Oson,O\'rta,Qiyin"', allow_blank=True)
    ws.add_data_validation(question_type_validation)
    ws.add_data_validation(answer_validation)
    ws.add_data_validation(difficulty_validation)
    question_type_validation.add("B2:B500")
    answer_validation.add("H2:H500")
    difficulty_validation.add("K2:K500")


def build_workbook():
    validate_banks()
    wb = Workbook()
    wb.remove(wb.active)

    settings = wb.create_sheet("Yo'nalish sozlamalari")
    settings.append(SETTINGS_HEADERS)
    for direction, rows in BANKS.items():
        settings.append([direction, sum(row["score"] for row in rows), 60, "Junior daraja: 20 Oson, 8 O'rta, 2 Qiyin"])
    style_header(settings, "548235")
    for col, width in {"A": 24, "B": 14, "C": 14, "D": 45}.items():
        settings.column_dimensions[col].width = width
    settings.freeze_panes = "A2"
    settings.auto_filter.ref = settings.dimensions

    widths = {"A": 8, "B": 14, "C": 55, "D": 32, "E": 32, "F": 32, "G": 32, "H": 14, "I": 70, "J": 12, "K": 18, "L": 24}
    for direction, rows in BANKS.items():
        ws = wb.create_sheet(direction)
        ws.append(HEADERS)
        for index, row in enumerate(rows, start=1):
            ws.append([
                index,
                row["type"],
                row["question"],
                row["a"],
                row["b"],
                row["c"],
                row["d"],
                row["correct"],
                row["rubric"],
                row["score"],
                row["difficulty"],
                row["topic"],
            ])
        style_header(ws, "1F4E78")
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        add_validations(ws)

    wb.save(OUTPUT)


def validate_workbook():
    wb = load_workbook(OUTPUT, read_only=True, data_only=True)
    expected = ["Yo'nalish sozlamalari", *BANKS.keys()]
    assert wb.sheetnames == expected, wb.sheetnames
    for direction in BANKS:
        ws = wb[direction]
        assert ws.max_row == 31, (direction, ws.max_row)
        assert ws.max_column == 12, (direction, ws.max_column)
    wb.close()


if __name__ == "__main__":
    build_workbook()
    validate_workbook()
    print(OUTPUT.resolve())
