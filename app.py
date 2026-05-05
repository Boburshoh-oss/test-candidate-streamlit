from __future__ import annotations

import json
import os
import random
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from streamlit.runtime.scriptrunner import get_script_run_ctx


load_dotenv()

QUESTIONS_FILE = Path("Junior_Yonalishlar_Savollar_30tadan.xlsx")
RESULTS_FILE = Path("candidate_results.xlsx")
ADMIN_PASSWORD = os.getenv("TEST_APP_ADMIN_PASSWORD", "admin123")
DEFAULT_MODEL = "gpt-5-mini"
SETTINGS_SHEET = "Yo'nalish sozlamalari"
DEFAULT_PASS_PERCENT = 60.0

STANDARD_HEADERS = [
    "No",
    "Savol turi",
    "Savol",
    "A variant",
    "B variant",
    "C variant",
    "D variant",
    "To'g'ri javob",
    "Kutilgan javob/Rubrika",
    "Maks ball",
    "Qiyinlik darajasi",
    "Mavzu",
]

RESULT_HEADERS = [
    "Yakunlangan vaqt",
    "Ism",
    "Telefon",
    "Yo'nalish",
    "Yo'nalish jami ball",
    "Jami maksimal ball",
    "Olingan ball",
    "Foiz",
    "O'tish foizi",
    "O'tish holati",
    "Baho",
    "Sarflangan vaqt (soniya)",
    "Jami savol",
    "To'g'ri test javoblari",
    "Noto'g'ri test javoblari",
    "Review kutayotgan yozma javoblar",
    "Javoblar JSON",
    "AI feedback JSON",
    "Review holati",
]


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = {
        "‘": "'",
        "’": "'",
        "`": "'",
        "ʻ": "'",
        "№": "no",
        " ": "",
        "_": "",
        "-": "",
        "/": "",
        "(": "",
        ")": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def get_secret(name: str) -> str | None:
    value = os.getenv(name)
    if value:
        return value
    try:
        return st.secrets.get(name)  # type: ignore[no-any-return]
    except Exception:
        return None


def first_existing(row: pd.Series, *names: str, default: Any = None) -> Any:
    normalized = {normalize_name(col): col for col in row.index}
    for name in names:
        key = normalize_name(name)
        if key in normalized:
            value = row.get(normalized[key])
            if pd.notna(value):
                return value
    return default


def parse_float(value: Any, default: float = 1.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def load_explanations(path_text: str) -> dict[str, str]:
    path = Path(path_text)
    if not path.exists():
        return {}
    try:
        df = pd.read_excel(path, sheet_name="Batafsil Javoblar")
    except Exception:
        return {}

    explanations: dict[str, str] = {}
    for _, row in df.iterrows():
        number = safe_text(first_existing(row, "No", "№", "Raqam"))
        question = safe_text(first_existing(row, "Savol"))
        explanation = safe_text(
            first_existing(
                row,
                "Yozma izoh (Tushuntirish)",
                "Izoh",
                "Tushuntirish",
                default="",
            )
        )
        if explanation:
            if number:
                explanations[f"no:{number}"] = explanation
            if question:
                explanations[f"q:{question}"] = explanation
    return explanations


def has_test_variants(columns: list[str]) -> bool:
    normalized = {normalize_name(col) for col in columns}
    return (
        normalize_name("A variant") in normalized
        and normalize_name("B variant") in normalized
        and normalize_name("C variant") in normalized
        and normalize_name("D variant") in normalized
    )


def is_question_sheet(sheet_name: str, df: pd.DataFrame) -> bool:
    if sheet_name in {"Baholash", "Batafsil Javoblar", SETTINGS_SHEET}:
        return False
    columns = list(df.columns)
    normalized = {normalize_name(col) for col in columns}
    has_question = normalize_name("Savol") in normalized
    has_type = normalize_name("Savol turi") in normalized
    return bool(has_question and (has_type or has_test_variants(columns)))


def load_question_bank(path_text: str) -> dict[str, list[dict[str, Any]]]:
    path = Path(path_text)
    if not path.exists():
        return {}

    workbook = pd.read_excel(path, sheet_name=None)
    explanations = load_explanations(path_text)
    directions: dict[str, list[dict[str, Any]]] = {}

    for sheet_name, df in workbook.items():
        if not is_question_sheet(sheet_name, df):
            continue

        questions: list[dict[str, Any]] = []
        for index, row in df.iterrows():
            question_text = safe_text(first_existing(row, "Savol"))
            if not question_text:
                continue

            number = safe_text(first_existing(row, "No", "№", "Raqam", default=index + 1))
            question_type = safe_text(first_existing(row, "Savol turi", "Turi", default=""))
            variants = {
                "A": safe_text(first_existing(row, "A variant", "A", default="")),
                "B": safe_text(first_existing(row, "B variant", "B", default="")),
                "C": safe_text(first_existing(row, "C variant", "C", default="")),
                "D": safe_text(first_existing(row, "D variant", "D", default="")),
            }
            if not question_type:
                question_type = "test" if all(variants.values()) else "yozma"

            question_type = question_type.lower().strip()
            if question_type not in {"test", "yozma"}:
                question_type = "test" if all(variants.values()) else "yozma"

            correct_answer = safe_text(
                first_existing(row, "To'g'ri javob", "Tog'ri javob", "To‘g‘ri javob", default="")
            )
            rubric = safe_text(
                first_existing(
                    row,
                    "Kutilgan javob/Rubrika",
                    "Rubrika",
                    "Kutilgan javob",
                    default="",
                )
            )
            explanation = explanations.get(f"no:{number}") or explanations.get(f"q:{question_text}") or ""
            if question_type == "yozma" and not rubric:
                rubric = correct_answer or explanation

            questions.append(
                {
                    "id": f"{sheet_name}-{number}-{index}",
                    "number": number,
                    "type": question_type,
                    "question": question_text,
                    "variants": variants,
                    "correct_answer": correct_answer.upper().strip()[:1],
                    "rubric": rubric,
                    "max_score": parse_float(first_existing(row, "Maks ball", "Max ball", default=1), 1),
                    "difficulty": safe_text(first_existing(row, "Qiyinlik darajasi", default="")),
                    "topic": safe_text(first_existing(row, "Mavzu", default="")),
                    "explanation": explanation,
                }
            )

        if questions:
            directions[sheet_name] = questions

    return directions


def load_direction_settings(path_text: str) -> dict[str, dict[str, float]]:
    path = Path(path_text)
    if not path.exists():
        return {}
    try:
        df = pd.read_excel(path, sheet_name=SETTINGS_SHEET)
    except Exception:
        return {}

    settings: dict[str, dict[str, float]] = {}
    for _, row in df.iterrows():
        direction = safe_text(first_existing(row, "Yo'nalish", "Yonalish", "Direction", default=""))
        if not direction:
            continue
        settings[direction] = {
            "declared_total": parse_float(first_existing(row, "Jami ball", "Jami maksimal ball", default=0), 0),
            "pass_percent": parse_float(first_existing(row, "O'tish foizi", "Otish foizi", default=DEFAULT_PASS_PERCENT), DEFAULT_PASS_PERCENT),
        }
    return settings


def get_direction_pass_percent(direction: str) -> float:
    settings = load_direction_settings(str(QUESTIONS_FILE))
    return settings.get(direction, {}).get("pass_percent", DEFAULT_PASS_PERCENT)


def get_direction_total_score(direction: str) -> float:
    bank = load_question_bank(str(QUESTIONS_FILE))
    questions = bank.get(direction, [])
    return sum(float(question.get("max_score", 0)) for question in questions)


def clear_question_cache() -> None:
    return None


def clamp_score(score: float, max_score: float) -> float:
    return max(0.0, min(float(score), float(max_score)))


def format_score(score: float) -> str:
    if abs(score - round(score)) < 0.0001:
        return str(int(round(score)))
    return f"{score:.2f}".rstrip("0").rstrip(".")


def get_grade(percent: float) -> str:
    if percent >= 90:
        return "A'lo - Junior+ darajada"
    if percent >= 75:
        return "Yaxshi - Junior darajaga mos"
    if percent >= 60:
        return "Qoniqarli - Trainee+ daraja"
    if percent >= 45:
        return "O'rtacha - qo'shimcha o'qish kerak"
    return "Yetarli emas - asoslarni qaytarish kerak"


def get_openai_client() -> Any | None:
    api_key = get_secret("OPENAI_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI

        return OpenAI(api_key=api_key)
    except Exception:
        return None


def grade_written_answer(
    *,
    question: str,
    candidate_answer: str,
    rubric: str,
    max_score: float,
    model: str,
) -> dict[str, Any]:
    if not candidate_answer.strip():
        return {
            "score": 0.0,
            "max_score": max_score,
            "feedback": "Yozma javob kiritilmagan.",
            "review_status": "empty_answer",
        }

    client = get_openai_client()
    if client is None:
        return {
            "score": None,
            "max_score": max_score,
            "feedback": "OpenAI API kaliti topilmadi yoki SDK ishga tushmadi. Qo'lda review kerak.",
            "review_status": "pending_manual_review",
        }

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "score": {"type": "number"},
            "max_score": {"type": "number"},
            "feedback": {"type": "string"},
            "review_status": {
                "type": "string",
                "enum": ["graded", "needs_manual_review"],
            },
        },
        "required": ["score", "max_score", "feedback", "review_status"],
    }

    prompt_payload = {
        "question": question,
        "rubric": rubric,
        "candidate_answer": candidate_answer,
        "max_score": max_score,
    }

    try:
        response = client.responses.create(
            model=model,
            input=[
                {
                    "role": "system",
                    "content": (
                        "Siz Uzbek tilidagi kandidat javoblarini baholaydigan tekshiruvchisiz. "
                        "Faqat berilgan savol va rubrikaga asoslaning. Javobni adolatli baholang, "
                        "ballni 0 va max_score oralig'ida bering, qisqa feedback yozing."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(prompt_payload, ensure_ascii=False),
                },
            ],
            text={
                "format": {
                    "type": "json_schema",
                    "name": "written_answer_grade",
                    "strict": True,
                    "schema": schema,
                }
            },
        )
        data = json.loads(response.output_text)
        return {
            "score": clamp_score(float(data["score"]), max_score),
            "max_score": max_score,
            "feedback": str(data["feedback"]),
            "review_status": str(data["review_status"]),
        }
    except Exception as exc:
        return {
            "score": None,
            "max_score": max_score,
            "feedback": f"AI baholashda xatolik: {exc}. Qo'lda review kerak.",
            "review_status": "pending_manual_review",
        }


def summarize_review_status(details: list[dict[str, Any]]) -> str:
    statuses = sorted({safe_text(item.get("review_status")) for item in details if item.get("review_status")})
    return ", ".join(statuses) if statuses else "graded"


def ensure_result_headers(ws: Any) -> list[str]:
    existing_headers = [safe_text(cell.value) for cell in ws[1]]
    if not any(existing_headers):
        for column_index, header in enumerate(RESULT_HEADERS, start=1):
            ws.cell(row=1, column=column_index, value=header)
        return RESULT_HEADERS[:]

    normalized_existing = {normalize_name(header): index for index, header in enumerate(existing_headers, start=1) if header}
    next_column = len(existing_headers) + 1
    for header in RESULT_HEADERS:
        if normalize_name(header) not in normalized_existing:
            ws.cell(row=1, column=next_column, value=header)
            existing_headers.append(header)
            next_column += 1
    return [safe_text(cell.value) for cell in ws[1]]


def append_result(result: dict[str, Any]) -> None:
    if RESULTS_FILE.exists():
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Natijalar"

    headers = ensure_result_headers(ws)
    ws.append([result.get(header, "") for header in headers])
    wb.save(RESULTS_FILE)


def ensure_workbook_exists() -> None:
    if QUESTIONS_FILE.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Junior"
    ws.append(STANDARD_HEADERS)
    wb.save(QUESTIONS_FILE)


def ensure_sheet_headers(ws: Any) -> list[str]:
    existing = [cell.value for cell in ws[1]]
    if not any(existing):
        for col_index, header in enumerate(STANDARD_HEADERS, start=1):
            ws.cell(row=1, column=col_index, value=header)
        return STANDARD_HEADERS[:]

    normalized_existing = {normalize_name(header): index for index, header in enumerate(existing, start=1)}
    headers = [safe_text(header) for header in existing if safe_text(header)]
    next_column = len(existing) + 1
    for header in STANDARD_HEADERS:
        if normalize_name(header) not in normalized_existing:
            ws.cell(row=1, column=next_column, value=header)
            headers.append(header)
            next_column += 1
    return [safe_text(cell.value) for cell in ws[1]]


def next_question_number(ws: Any) -> int:
    number_col = None
    for index, cell in enumerate(ws[1], start=1):
        if normalize_name(cell.value) in {normalize_name("No"), "no", "raqam"}:
            number_col = index
            break
    if number_col is None:
        return max(1, ws.max_row)

    max_number = 0
    for row in range(2, ws.max_row + 1):
        try:
            max_number = max(max_number, int(ws.cell(row=row, column=number_col).value or 0))
        except Exception:
            continue
    return max_number + 1


def append_question_to_workbook(sheet_name: str, values: dict[str, Any]) -> None:
    ensure_workbook_exists()
    wb = load_workbook(QUESTIONS_FILE)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    headers = ensure_sheet_headers(ws)
    values["№"] = next_question_number(ws)

    row_index = ws.max_row + 1
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=row_index, column=column_index, value=values.get(header, ""))
    wb.save(QUESTIONS_FILE)
    clear_question_cache()


def save_uploaded_workbook(uploaded_file: Any) -> None:
    QUESTIONS_FILE.write_bytes(uploaded_file.getvalue())
    clear_question_cache()


def read_results_file() -> bytes | None:
    if not RESULTS_FILE.exists():
        return None
    return RESULTS_FILE.read_bytes()


def initialize_state() -> None:
    defaults = {
        "stage": "registration",
        "admin_authenticated": False,
        "time_limit_minutes": 20,
        "question_limit": 0,
        "openai_model": DEFAULT_MODEL,
        "candidate": {},
        "direction": "",
        "questions": [],
        "started_at": None,
        "ends_at": None,
        "result": None,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)


def reset_candidate_flow() -> None:
    for key in list(st.session_state.keys()):
        if key.startswith("answer_"):
            del st.session_state[key]
    st.session_state.stage = "registration"
    st.session_state.candidate = {}
    st.session_state.direction = ""
    st.session_state.questions = []
    st.session_state.started_at = None
    st.session_state.ends_at = None
    st.session_state.result = None


def render_admin_panel(directions: dict[str, list[dict[str, Any]]]) -> None:
    st.sidebar.header("Admin")
    if not st.session_state.admin_authenticated:
        password = st.sidebar.text_input("Admin parol", type="password")
        if st.sidebar.button("Kirish"):
            if password == ADMIN_PASSWORD:
                st.session_state.admin_authenticated = True
                st.rerun()
            st.sidebar.error("Parol noto'g'ri")
        return

    st.sidebar.success("Admin panel ochiq")
    if st.sidebar.button("Chiqish"):
        st.session_state.admin_authenticated = False
        st.rerun()

    st.sidebar.number_input(
        "Test vaqti (daqiqa)",
        min_value=1,
        max_value=240,
        value=int(st.session_state.time_limit_minutes),
        key="time_limit_minutes",
    )
    st.sidebar.number_input(
        "Savollar soni (0 = hammasi)",
        min_value=0,
        max_value=500,
        value=int(st.session_state.question_limit),
        key="question_limit",
    )
    st.sidebar.text_input("OpenAI model", key="openai_model")

    st.sidebar.divider()
    st.sidebar.subheader("Natijalar")
    results_bytes = read_results_file()
    if results_bytes:
        st.sidebar.caption(f"Fayl: {RESULTS_FILE.name} ({len(results_bytes) / 1024:.1f} KB)")
        st.sidebar.download_button(
            "Natijalarni yuklab olish",
            data=results_bytes,
            file_name=RESULTS_FILE.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.sidebar.caption("Hali natija fayli yaratilmagan.")

    st.sidebar.divider()
    uploaded_file = st.sidebar.file_uploader("Yangi Excel workbook yuklash", type=["xlsx"])
    if uploaded_file and st.sidebar.button("Workbookni saqlash"):
        try:
            save_uploaded_workbook(uploaded_file)
            st.sidebar.success("Workbook saqlandi")
            st.rerun()
        except Exception as exc:
            st.sidebar.error(f"Workbook saqlanmadi: {exc}")

    st.sidebar.divider()
    with st.sidebar.expander("Savol qo'shish", expanded=False):
        existing_directions = sorted(directions.keys())
        target_mode = st.radio("Yo'nalish", ["Mavjud", "Yangi"], horizontal=True)
        if target_mode == "Mavjud" and existing_directions:
            sheet_name = st.selectbox("Mavjud yo'nalish", existing_directions)
        else:
            sheet_name = st.text_input("Yangi yo'nalish nomi", value="")

        with st.form("add_question_form", clear_on_submit=True):
            question_type = st.selectbox("Savol turi", ["test", "yozma"])
            question = st.text_area("Savol")
            max_score = st.number_input("Maks ball", min_value=0.1, max_value=100.0, value=1.0, step=0.5)
            topic = st.text_input("Mavzu")
            difficulty = st.selectbox("Qiyinlik darajasi", ["", "Oson", "O'rta", "Qiyin"])
            a_variant = b_variant = c_variant = d_variant = correct_answer = rubric = ""
            if question_type == "test":
                a_variant = st.text_input("A variant")
                b_variant = st.text_input("B variant")
                c_variant = st.text_input("C variant")
                d_variant = st.text_input("D variant")
                correct_answer = st.selectbox("To'g'ri javob", ["A", "B", "C", "D"])
            else:
                rubric = st.text_area("Kutilgan javob/Rubrika")

            submitted = st.form_submit_button("Savolni qo'shish")
            if submitted:
                if not sheet_name.strip() or not question.strip():
                    st.error("Yo'nalish nomi va savol majburiy.")
                else:
                    append_question_to_workbook(
                        sheet_name.strip(),
                        {
                            "Savol turi": question_type,
                            "Savol": question.strip(),
                            "A variant": a_variant.strip(),
                            "B variant": b_variant.strip(),
                            "C variant": c_variant.strip(),
                            "D variant": d_variant.strip(),
                            "To'g'ri javob": correct_answer,
                            "Kutilgan javob/Rubrika": rubric.strip(),
                            "Maks ball": max_score,
                            "Qiyinlik darajasi": difficulty,
                            "Mavzu": topic.strip(),
                        },
                    )
                    st.success("Savol qo'shildi.")


def choose_questions(all_questions: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
    if limit <= 0 or limit >= len(all_questions):
        return all_questions[:]
    return random.sample(all_questions, limit)


def render_registration(directions: dict[str, list[dict[str, Any]]]) -> None:
    st.title("Ishonch IT Department Test Platformasi")
    st.caption("Ishonch IT Department uchun kandidat baholash tizimi.")

    if not directions:
        st.warning("Savollar topilmadi. Admin panel orqali Excel workbook yuklang yoki savol qo'shing.")
        return

    with st.form("registration_form"):
        name = st.text_input("Ism va familiya")
        phone = st.text_input("Telefon")
        direction = st.selectbox("Yo'nalish", sorted(directions.keys()))
        submitted = st.form_submit_button("Testni boshlash")

    if submitted:
        if not name.strip() or not phone.strip():
            st.error("Ism va telefon majburiy.")
            return

        questions = choose_questions(directions[direction], int(st.session_state.question_limit))
        if not questions:
            st.error("Tanlangan yo'nalishda savollar yo'q.")
            return

        st.session_state.candidate = {"name": name.strip(), "phone": phone.strip()}
        st.session_state.direction = direction
        st.session_state.questions = questions
        now = time.time()
        st.session_state.started_at = now
        st.session_state.ends_at = now + int(st.session_state.time_limit_minutes) * 60
        st.session_state.result = None
        st.session_state.stage = "test"
        st.rerun()


def collect_answers() -> dict[str, str]:
    answers: dict[str, str] = {}
    for question in st.session_state.questions:
        key = f"answer_{question['id']}"
        answers[question["id"]] = safe_text(st.session_state.get(key, ""))
    return answers


def finish_test() -> None:
    questions = st.session_state.questions
    answers = collect_answers()
    started_at = float(st.session_state.started_at or time.time())
    spent_seconds = int(time.time() - started_at)

    total_possible = sum(float(question["max_score"]) for question in questions)
    earned = 0.0
    test_correct = 0
    test_wrong = 0
    pending_review = 0
    details: list[dict[str, Any]] = []
    ai_feedback: list[dict[str, Any]] = []

    with st.spinner("Javoblar baholanmoqda..."):
        for question in questions:
            answer = answers.get(question["id"], "")
            max_score = float(question["max_score"])
            if question["type"] == "test":
                is_correct = answer.upper() == question["correct_answer"].upper()
                score = max_score if is_correct else 0.0
                earned += score
                test_correct += int(is_correct)
                test_wrong += int(not is_correct)
                review_status = "graded"
                feedback = question.get("explanation") or (
                    "To'g'ri javob." if is_correct else f"To'g'ri javob: {question['correct_answer']}"
                )
            else:
                graded = grade_written_answer(
                    question=question["question"],
                    candidate_answer=answer,
                    rubric=question["rubric"],
                    max_score=max_score,
                    model=st.session_state.openai_model,
                )
                score_value = graded.get("score")
                if score_value is None:
                    score = 0.0
                    pending_review += 1
                else:
                    score = clamp_score(float(score_value), max_score)
                    earned += score
                review_status = graded["review_status"]
                feedback = graded["feedback"]
                ai_feedback.append(
                    {
                        "number": question["number"],
                        "question": question["question"],
                        "score": score_value,
                        "max_score": max_score,
                        "feedback": feedback,
                        "review_status": review_status,
                    }
                )

            details.append(
                {
                    "number": question["number"],
                    "type": question["type"],
                    "question": question["question"],
                    "candidate_answer": answer,
                    "correct_answer_or_rubric": question["correct_answer"] if question["type"] == "test" else question["rubric"],
                    "score": score if question["type"] == "test" or review_status != "pending_manual_review" else None,
                    "max_score": max_score,
                    "feedback": feedback,
                    "review_status": review_status,
                }
            )

    percent = (earned / total_possible * 100) if total_possible else 0
    pass_percent = get_direction_pass_percent(st.session_state.direction)
    pass_status = "O'tdi" if percent >= pass_percent else "O'tmadi"
    direction_total = get_direction_total_score(st.session_state.direction)
    result = {
        "Yakunlangan vaqt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Ism": st.session_state.candidate.get("name", ""),
        "Telefon": st.session_state.candidate.get("phone", ""),
        "Yo'nalish": st.session_state.direction,
        "Yo'nalish jami ball": format_score(direction_total),
        "Jami maksimal ball": format_score(total_possible),
        "Olingan ball": format_score(earned),
        "Foiz": round(percent, 2),
        "O'tish foizi": round(pass_percent, 2),
        "O'tish holati": pass_status,
        "Baho": get_grade(percent),
        "Sarflangan vaqt (soniya)": spent_seconds,
        "Jami savol": len(questions),
        "To'g'ri test javoblari": test_correct,
        "Noto'g'ri test javoblari": test_wrong,
        "Review kutayotgan yozma javoblar": pending_review,
        "Javoblar JSON": json.dumps(details, ensure_ascii=False),
        "AI feedback JSON": json.dumps(ai_feedback, ensure_ascii=False),
        "Review holati": summarize_review_status(details),
        "_details": details,
    }
    append_result(result)
    st.session_state.result = result
    st.session_state.stage = "result"


def render_timer() -> int:
    ends_at = float(st.session_state.ends_at or time.time())
    remaining = max(0, int(ends_at - time.time()))
    minutes, seconds = divmod(remaining, 60)
    st.metric("Qolgan vaqt", f"{minutes:02d}:{seconds:02d}")
    if st.session_state.time_limit_minutes:
        total_seconds = int(st.session_state.time_limit_minutes) * 60
        progress = 1 - min(remaining / total_seconds, 1)
        st.progress(progress)
    return remaining


def render_question(question: dict[str, Any], display_index: int) -> None:
    title = f"{display_index}. {question['question']}"
    meta = []
    if question.get("topic"):
        meta.append(question["topic"])
    if question.get("difficulty"):
        meta.append(question["difficulty"])
    meta_text = f" ({' / '.join(meta)})" if meta else ""
    st.subheader(title + meta_text)
    st.caption(f"Maks ball: {format_score(float(question['max_score']))}")

    key = f"answer_{question['id']}"
    if question["type"] == "test":
        options = ["A", "B", "C", "D"]

        def option_label(option: str) -> str:
            text = question["variants"].get(option, "")
            return f"{option}. {text}" if text else option

        st.radio(
            "Javobni tanlang",
            options,
            key=key,
            format_func=option_label,
            index=None,
        )
    else:
        st.text_area("Yozma javob", key=key, height=140)


def render_test() -> None:
    st.title(st.session_state.direction)
    candidate = st.session_state.candidate
    st.caption(f"Kandidat: {candidate.get('name', '')} | Telefon: {candidate.get('phone', '')}")

    remaining = render_timer()
    if remaining <= 0:
        st.warning("Vaqt tugadi. Javoblar yakunlanmoqda.")
        finish_test()
        st.rerun()

    st.divider()
    for index, question in enumerate(st.session_state.questions, start=1):
        render_question(question, index)
        st.divider()

    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("Testni yakunlash", type="primary"):
            finish_test()
            st.rerun()
    with col2:
        answered = sum(1 for value in collect_answers().values() if value)
        st.caption(f"Javob berilgan: {answered}/{len(st.session_state.questions)}")

    time.sleep(1)
    st.rerun()


def render_result() -> None:
    result = st.session_state.result
    if not result:
        st.session_state.stage = "registration"
        st.rerun()

    st.title("Test natijasi")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Ball", f"{result['Olingan ball']} / {result['Jami maksimal ball']}")
    col2.metric("Foiz", f"{result['Foiz']}%")
    pass_percent_value = result.get("O'tish foizi", DEFAULT_PASS_PERCENT)
    col3.metric("O'tish foizi", f"{pass_percent_value}%")
    col4.metric("Holat", result.get("O'tish holati", ""))

    st.caption(f"Baho: {result['Baho']}")

    if result["Review kutayotgan yozma javoblar"]:
        st.warning(
            "Ba'zi yozma javoblar AI orqali baholanmadi. Ular natija Excel faylida review uchun saqlandi."
        )
    else:
        st.success("Natija saqlandi.")

    st.subheader("Javoblar tahlili")
    for item in result["_details"]:
        with st.expander(f"{item['number']}. {item['question']}"):
            st.write(f"Javob: {item['candidate_answer'] or 'Javob berilmagan'}")
            if item["type"] == "test":
                st.write(f"To'g'ri javob: {item['correct_answer_or_rubric']}")
            else:
                st.write(f"Rubrika: {item['correct_answer_or_rubric'] or 'Kiritilmagan'}")
            score_text = "review kutmoqda" if item["score"] is None else format_score(float(item["score"]))
            st.write(f"Ball: {score_text} / {format_score(float(item['max_score']))}")
            st.write(f"Izoh: {item['feedback']}")

    if st.button("Yangi kandidat", type="primary"):
        reset_candidate_flow()
        st.rerun()


def main() -> None:
    st.set_page_config(page_title="Ishonch IT Department Test Platformasi", page_icon="Test", layout="wide")
    initialize_state()
    ensure_workbook_exists()
    directions = load_question_bank(str(QUESTIONS_FILE))
    render_admin_panel(directions)

    if st.session_state.stage == "registration":
        render_registration(directions)
    elif st.session_state.stage == "test":
        render_test()
    elif st.session_state.stage == "result":
        render_result()
    else:
        reset_candidate_flow()
        render_registration(directions)


if __name__ == "__main__":
    if get_script_run_ctx() is None:
        print("Streamlit ilova ishga tushirilmoqda: streamlit run app.py")
        subprocess.run([sys.executable, "-m", "streamlit", "run", __file__], check=False)
    else:
        main()
